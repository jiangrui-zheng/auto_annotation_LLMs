import pandas as pd
from openpyxl import load_workbook


def process_trailing(this_sent):
    this_sent = this_sent.replace('"', "'")
    this_sent = this_sent.replace('’', "'")
    this_sent = this_sent.replace('”', "'").replace('“', "'")
    this_sent = ' '.join(this_sent.split())
    return this_sent


label_description_dict = {
    1: ("1: Non-hateful Counter Hate Speech", "Respectful disagreement to hate speech"),
    2: ("2: Aggressive/Hateful Counter Hate Speech", "Including attack to people who are hateful towards protected group"),
    3: ("3: Untargeted Hate Speech", "General derogatory language, not group-specific"),
    4: ("4: Non-hateful Speech Discussing Hate", "Critique, facts, or asking a questions without promoting hate"),
    5: ("5: Paraphrasing or Quoting Hate Speech", "Describing third person/group's words without promoting"),
    6: ("6: Supporting Protected Group", "expresses solidarity, support, or encouragement"),
    7: ("7: Attacking Non-Protected Group", "such as organization, objects, politicians, nazis, gangsters, criminals"),
    8: ("8: Actual hate speech", "Attacking Protected Group")
}

file_name = 'Hatemoderate Hard-negative.xlsx'
book = load_workbook(file_name)

data1 = []
data2 = []


for sheet in book.sheetnames[10:17]:
    df = pd.read_excel(file_name, sheet_name=sheet, usecols="B:C", engine='openpyxl')
    row_start_at = df[df.iloc[:, 0] == 'non-hate examples'].index[0] + 1
    df = df.iloc[row_start_at:][:101]
    df = df.dropna()
    df = df[df.iloc[:, 1].apply(lambda x: isinstance(x, (int, float)))]

    for _, row in df.iterrows():
        try:
            sent = process_trailing(row.iloc[0])
            label_num = row.iloc[1]
            #if pd.isnull(label_num) or not isinstance(label_num, (int, float)):
                #label_num = 8
            label, description = label_description_dict[label_num]
            data1.append("{\"prompt\":\"" + sent + "\", \"completion\": \"" + label + "\"}\n")
            data2.append("{\"prompt\":\"" + sent + "\", \"completion\": \"" + label + ": " + process_trailing(description) + "\"}\n")
        except KeyError:
            print(print(row.iloc[0]))


extra_df = pd.read_csv("../postprocess/all_examples_0601_hate.csv", sep='\t')
#print(extra_df.columns)

grouped = extra_df.groupby("guideline")
selected_sentences = grouped.apply(lambda x: x.sample(n=3, random_state=1) if len(x) > 3 else x)

for _, row in selected_sentences.iterrows():
    sent = process_trailing(row["sentence"])
    label_num = 8
    label, description = label_description_dict[label_num]
    data1.append("{\"prompt\":\"" + sent + "\", \"completion\": \"" + label + "\"}\n")
    data2.append("{\"prompt\":\"" + sent + "\", \"completion\": \"" + label + ": " + process_trailing(description) + "\"}\n")


with open('train_pr0.jsonl', 'w') as f:
    for entry in data1:
        f.write(entry)

with open('train_pr1.jsonl', 'w') as f:
    for entry in data2:
        f.write(entry)
